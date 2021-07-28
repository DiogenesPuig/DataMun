#!/usr/bin/env python

import string
from .models import *
import math
import random
from time import time
import sqlite3
import numpy as np
import openpyxl
import psycopg2
from django.conf import settings

def workbookToSqlStatements(workbook,table_name,sheet_name_0,sheet_name_splitter,sheet_name_1,cols_name,ints,length_varchar,min_row):
    """
    Returns create and insert sql statement's from a given workbook 
    """
    create_sql_statement = ""
    create_sql_statement = f"CREATE TABLE IF NOT EXISTS {table_name} "
    create_sql_statement += f"(id SERIAL PRIMARY KEY , {sheet_name_0} INT, {sheet_name_1} INT,"
    cols = str(cols_name).strip("]")
    cols = cols.strip("[")
    for i in range(len(cols_name)):
        cols = cols.replace("'","")
        cols = cols.replace("'","")
        if i in ints:
            create_sql_statement += f"{cols_name[i]} INT"
        else:
            create_sql_statement += f"{cols_name[i]} CHAR({length_varchar})"

        if i != len(cols_name) - 1:
            create_sql_statement += ","
    create_sql_statement += ");"


    insert_sql_statement = ""
    insert_sql_statement = f"insert into {table_name} ({sheet_name_0},{sheet_name_1},{cols}) values "
    for sheet in workbook.worksheets:
        sheet_name_list = str(sheet.title).split(sheet_name_splitter)
        year = sheet_name_list[0]
        week = sheet_name_list[1]
        for row in sheet.iter_rows(min_row=min_row,values_only=True):
            values = "("
            values += f"{year},{week},"
            insert = True
            for i in range(len(cols_name)):
                try:
                    if i in ints:

                        cell = str(row[i])
                        try:
                            values += f"{int(cell)}"
                        except:
                            values += "null"
                    else:
                        if str(row[i]) != "" and str(row[i]) != "None":
                            values += f"'{str(row[i]).rstrip()}'"
                        else:
                            insert = False

                    if i != len(cols_name) - 1:
                        values += ","
                    else:
                        break
                except:
                    return print("error the length of the list cols_name exeed the length of max columns of the sheet")

            values += ")"

            if insert:
                insert_sql_statement += values + ",\n"
    temp = len(insert_sql_statement)
    insert_sql_statement = insert_sql_statement[:temp - 2]
    insert_sql_statement += ";"

    #insert_sql_statement = insert_sql_statement.replace(",\n;","\n;")
    return create_sql_statement, insert_sql_statement

hostname = settings.DATABASES["default"]["HOST"]
username = settings.DATABASES["default"]["USER"]
password = settings.DATABASES["default"]["PASSWORD"]
database = settings.DATABASES["default"]["NAME"]


insert = """
--DELETE from dataMun_diagnostic ;

-- deleting cols useless
delete from raw where col0 is null;
-- sql datazone
insert into "dataMun_zone" (code) 
select DISTINCT col0 from raw where col0 not in (select code from "dataMun_zone");

--- center
insert into "dataMun_center" (code,name,zone_id) 
select DISTINCT col1, col2, (select dz.id from "dataMun_zone" dz where col0 = dz.code) 
from raw where col1 not in (select code from "dataMun_center");

-- diagnostic
insert into "dataMun_diagnostic" (code,name) 
select DISTINCT col3, col4 
from raw where col3 not in (select code from "dataMun_diagnostic");

-- year
insert into "dataMun_year" (year)
select distinct r.year from raw r
where r."year" not in(select "year" from "dataMun_year");


-- week
insert into dataMun_week (week,year_id,creation)
	select distinct r.week, 
	(SELECT dy.id from dataMun_year dy where dy.year = r.year),
	current_date
from raw r; 
--where r.week not in (select week from dataMun_week) and r.year not in (select year from dataMun_year);

-- 
insert into "dataMun_sex"(name) values ('Male');
insert into "dataMun_sex"(name) values ('Female');

insert into "dataMun_age" (from_age,to_age) values(0,1);
insert into "dataMun_age" (from_age,to_age) values(1,5);
insert into "dataMun_age" (from_age,to_age) values(6,9);
insert into "dataMun_age" (from_age,to_age) values(10,14);
insert into "dataMun_age" (from_age,to_age) values(15,19);
insert into "dataMun_age" (from_age,to_age) values(20,54);
insert into "dataMun_age" (from_age,to_age) values(55,65);
insert into "dataMun_age" (from_age,to_age) values(65,214748367);


-- diagnostic cases

select * from "dataMun_age" dma ;
-- cases <1 M
insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col5,
	(select dc.id from dataMun_center dc where r.col1 = dc.code ),
	(select dd.id from dataMun_diagnostic dd where r.col3 = dd.code),
	(select dw.id from dataMun_week dw where r.week = dw.week),
	(select ds.id from dataMun_sex ds where ds.name like 'M%'),
	(select da.id from dataMun_age da where da.from_age = 0)
from raw r where col5 is not null and col5 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id = (select da.id from dataMun_age da where da.from_age = 0))
;

-- cases <1 F
insert into "dataMun_diagnosticcases" (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col6,
	(select dc.id from "dataMun_center" dc where raw.col1 = dc.code ),
	(select dd.id from "dataMun_diagnostic" dd where raw.col3 = dd.code),
	(select dw.id from "dataMun_week" dw where raw.week = dw.week),
	(select ds.id from "dataMun_sex" ds where ds.name like 'F%'),
	(select da.id from "dataMun_age" da where da.from_age = 0)
from raw where col6 is not null and col6 not in (select cases from "dataMun_diagnosticcases" where sex_id=2 and age_id=(select da.id from "dataMun_age" da where da.from_age = 0))
;


-- cases 1 to 5 M
insert into "dataMun_diagnosticcases" (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col7,
	(select dc.id from "dataMun_center" dc where raw.col1 = dc.code ),
	(select dd.id from "dataMun_diagnostic" dd where raw.col3 = dd.code),
	(select dw.id from "dataMun_week" dw where raw.week = dw.week),
	(select ds.id from "dataMun_sex" ds where ds.name like 'M%'),
	(select da.id from "dataMun_age" da where da.from_age = 1)
from raw where col7 is not null and col7 not in (select cases from "dataMun_diagnosticcases" where sex_id=1 and age_id=(select da.id from "dataMun_age" da where da.from_age = 1))
;

select  * from "dataMun_age" dma where from_age = 1;
-- cases 1 to 5 F
insert into "dataMun_diagnosticcases" (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col8,
	(select dc.id from "dataMun_center" dc where raw.col1 = dc.code ),
	(select dd.id from "dataMun_diagnostic" dd where raw.col3 = dd.code),
	(select dw.id from "dataMun_week" dw where raw.week = dw.week),
	(select ds.id from "dataMun_sex" ds where ds.name like 'F%'),
	(select da.id from "dataMun_age" da where da.from_age = 1)
from raw where col8 is not null and col8 not in (select cases from "dataMun_diagnosticcases" where sex_id=2 and age_id=(select da.id from "dataMun_age" da where da.from_age = 1))
;

-- cases 6 to 9 M
insert into "dataMun_diagnosticcases" (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col9,
	(select dc.id from "dataMun_center" dc where raw.col1 = dc.code ),
	(select dd.id from "dataMun_diagnostic" dd where raw.col3 = dd.code),
	(select dw.id from "dataMun_week" dw where raw.week = dw.week),
	(select ds.id from "dataMun_sex" ds where ds.name like 'M%'),
	(select da.id from "dataMun_age" da where da.from_age = 6)
from raw where col9 is not null and col9 not in (select cases from "dataMun_diagnosticcases" where sex_id=1 and age_id = (select da.id from "dataMun_age" da where da.from_age = 6))
;

-- cases 6 to 9 F
insert into "dataMun_diagnosticcases" (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col10,
	(select dc.id from "dataMun_center" dc where raw.col1 = dc.code ),
	(select dd.id from "dataMun_diagnostic" dd where raw.col3 = dd.code),
	(select dw.id from "dataMun_week" dw where raw.week = dw.week),
	(select ds.id from "dataMun_sex" ds where ds.name like 'F%'),
	(select da.id from "dataMun_age" da where da.from_age = 6)
from raw where col10 is not null and col10 not in (select cases from "dataMun_diagnosticcases" where sex_id=2 and age_id=(select da.id from "dataMun_age" da where da.from_age = 6))
;
-- cases 10 to 14 M
insert into "dataMun_diagnosticcases" (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col11,
	(select dc.id from "dataMun_center" dc where raw.col1 = dc.code ),
	(select dd.id from "dataMun_diagnostic" dd where raw.col3 = dd.code),
	(select dw.id from "dataMun_week" dw where raw.week = dw.week),
	(select ds.id from "dataMun_sex" ds where ds.name like 'M%'),
	(select da.id from "dataMun_age" da where da.from_age = 10)
from raw where col11 is not null and col11 not in (select cases from "dataMun_diagnosticcases" where sex_id=1 and age_id=(select da.id from "dataMun_age" da where da.from_age = 10))
;
-- cases 10 to 14 F
insert into "dataMun_diagnosticcases" (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col12,
	(select dc.id from "dataMun_center" dc where raw.col1 = dc.code ),
	(select dd.id from "dataMun_diagnostic" dd where raw.col3 = dd.code),
	(select dw.id from "dataMun_week" dw where raw.week = dw.week),
	(select ds.id from "dataMun_sex" ds where ds.name like 'F%'),
	(select da.id from "dataMun_age" da where da.from_age = 10)
from raw where col12 is not null and col12 not in (select cases from "dataMun_diagnosticcases" where sex_id=2 and age_id=(select da.id from "dataMun_age" da where da.from_age = 10))
;

-- 15 to 19 m
insert into "dataMun_diagnosticcases" (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col13,
	(select dc.id from "dataMun_center" dc where raw.col1 = dc.code ),
	(select dd.id from "dataMun_diagnostic" dd where raw.col3 = dd.code),
	(select dw.id from "dataMun_week" dw where raw.week = dw.week),
	(select ds.id from "dataMun_sex" ds where ds.name like 'M%'),
	(select da.id from "dataMun_age" da where da.from_age = 15)
from raw where col13 is not null and col13 not in (select cases from "dataMun_diagnosticcases" where sex_id=1 and age_id=(select da.id from "dataMun_age" da where da.from_age = 15))
;

-- 15 to 19 f
insert into "dataMun_diagnosticcases" (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col14,
	(select dc.id from "dataMun_center" dc where raw.col1 = dc.code ),
	(select dd.id from "dataMun_diagnostic" dd where raw.col3 = dd.code),
	(select dw.id from "dataMun_week" dw where raw.week = dw.week),
	(select ds.id from "dataMun_sex" ds where ds.name like 'F%'),
	(select da.id from "dataMun_age" da where da.from_age = 15)
from raw where col14 is not null and col14 not in (select cases from "dataMun_diagnosticcases" where sex_id=2 and age_id=(select da.id from "dataMun_age" da where da.from_age = 15))
;

-- 20 to 54 m
insert into "dataMun_diagnosticcases" (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col15,
	(select dc.id from "dataMun_center" dc where raw.col1 = dc.code ),
	(select dd.id from "dataMun_diagnostic" dd where raw.col3 = dd.code),
	(select dw.id from "dataMun_week" dw where raw.week = dw.week),
	(select ds.id from "dataMun_sex" ds where ds.name like 'M%'),
	(select da.id from "dataMun_age" da where da.from_age = 20)
from raw where col15 is not null and col15 not in (select cases from "dataMun_diagnosticcases" where sex_id=1 and age_id=(select da.id from "dataMun_age" da where da.from_age = 20))
;

-- 20 to 54 f
insert into "dataMun_diagnosticcases" (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col16,
	(select dc.id from "dataMun_center" dc where raw.col1 = dc.code ),
	(select dd.id from "dataMun_diagnostic" dd where raw.col3 = dd.code),
	(select dw.id from "dataMun_week" dw where raw.week = dw.week),
	(select ds.id from "dataMun_sex" ds where ds.name like 'F%'),
	(select da.id from "dataMun_age" da where da.from_age = 20)
from raw where col16 is not null and col16 not in (select cases from "dataMun_diagnosticcases" where sex_id=2 and age_id=(select da.id from "dataMun_age" da where da.from_age = 20))
;

-- 55 to 65 m
insert into "dataMun_diagnosticcases" (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col17,
	(select dc.id from "dataMun_center" dc where raw.col1 = dc.code ),
	(select dd.id from "dataMun_diagnostic" dd where raw.col3 = dd.code),
	(select dw.id from "dataMun_week" dw where raw.week = dw.week),
	(select ds.id from "dataMun_sex" ds where ds.name like 'M%'),
	(select da.id from "dataMun_age" da where da.from_age = 55)
from raw where col17 is not null and col17 not in (select cases from "dataMun_diagnosticcases" where sex_id=1 and age_id=(select da.id from "dataMun_age" da where da.from_age = 55))
;

-- 55 to 65 f
insert into "dataMun_diagnosticcases" (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col18,
	(select dc.id from "dataMun_center" dc where raw.col1 = dc.code ),
	(select dd.id from "dataMun_diagnostic" dd where raw.col3 = dd.code),
	(select dw.id from "dataMun_week" dw where raw.week = dw.week),
	(select ds.id from "dataMun_sex" ds where ds.name like 'F%'),
	(select da.id from "dataMun_age" da where da.from_age = 55)
from raw where col18 is not null and col18 not in (select cases from "dataMun_diagnosticcases" where sex_id=2 and age_id=(select da.id from "dataMun_age" da where da.from_age = 55))
;
-- more than 65 m
insert into "dataMun_diagnosticcases" (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col19,
	(select dc.id from "dataMun_center" dc where raw.col1 = dc.code ),
	(select dd.id from "dataMun_diagnostic" dd where raw.col3 = dd.code),
	(select dw.id from "dataMun_week" dw where raw.week = dw.week),
	(select ds.id from "dataMun_sex" ds where ds.name like 'M%'),
	(select da.id from "dataMun_age" da where da.from_age = 65)
from raw where col19 is not null and col19 not in (select cases from "dataMun_diagnosticcases" where sex_id=1 and age_id=(select da.id from "dataMun_age" da where da.from_age = 65))
;

-- more than 65 f
insert into "dataMun_diagnosticcases" (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id)
select current_date,
	col20,
	(select dc.id from "dataMun_center" dc where raw.col1 = dc.code ),
	(select dd.id from "dataMun_diagnostic" dd where raw.col3 = dd.code),
	(select dw.id from "dataMun_week" dw where raw.week = dw.week),
	(select ds.id from "dataMun_sex" ds where ds.name like 'F%'),
	(select da.id from "dataMun_age" da where da.from_age = 65)
from raw where col20 is not null and col20 not in (select cases from "dataMun_diagnosticcases" where sex_id=2 and age_id=(select da.id from "dataMun_age" da where da.from_age = 65))
;

"""
def insertSQL():

    conn = sqlite3.connect('db.sqlite3')
    c = conn.cursor()

    c.execute("delete from raw where col0 is null;")
    c.execute("insert into dataMun_zone (code) select DISTINCT col0 from raw where col0 not in (select code from dataMun_zone);")
    c.execute("insert into dataMun_center (code,name,zone_id) select DISTINCT col1, col2, (select dz.id from dataMun_zone dz where col0 = dz.code) from raw where col1 not in (select code from dataMun_center);")
    c.execute("insert into dataMun_diagnostic (code,name) select DISTINCT col3, col4 from raw where col3 not in (select code from dataMun_diagnostic);")
    c.execute("insert into dataMun_year (year) select distinct r.year from raw r where r.year not in(select dy.year from dataMun_year dy);")
    c.execute("insert into dataMun_week (week,year_id,creation) select distinct r.week, (SELECT dy.id from dataMun_year dy where dy.year = r.year),current_date from raw r;")
    c.execute("insert into dataMun_sex(name) values ('Male'),('Female');")
    c.execute("insert into dataMun_age (from_age,to_age) values(0,1),(1,5),(6,9),(10,14),(15,19),(20,54),(55,65),(65,214748367);")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date,col5,(select dc.id from dataMun_center dc where r.col1 = dc.code ),(select dd.id from dataMun_diagnostic dd where r.col3 = dd.code),(select dw.id from dataMun_week dw where r.week = dw.week),(select ds.id from dataMun_sex ds where ds.name like 'M%'),(select da.id from dataMun_age da where da.from_age = 0) from raw r where col5 is not null and col5 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id = (select da.id from dataMun_age da where da.from_age = 0));")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col6, (select dc.id from dataMun_center dc where raw.col1 = dc.code ),(select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code),(select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'F%'), (select da.id from dataMun_age da where da.from_age = 0) from raw where col6 is not null and col6 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 0));")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date,col7, (select dc.id from dataMun_center dc where raw.col1 = dc.code ),	(select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code),	(select dw.id from dataMun_week dw where raw.week = dw.week),(select ds.id from dataMun_sex ds where ds.name like 'M%'),	(select da.id from dataMun_age da where da.from_age = 1) from raw where col7 is not null and col7 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id=(select da.id from dataMun_age da where da.from_age = 1));")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date,col8, (select dc.id from dataMun_center dc where raw.col1 = dc.code ),	(select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code),	(select dw.id from dataMun_week dw where raw.week = dw.week),	(select ds.id from dataMun_sex ds where ds.name like 'F%'),	(select da.id from dataMun_age da where da.from_age = 1) from raw where col8 is not null and col8 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 1));")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col9, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'M%'), (select da.id from dataMun_age da where da.from_age = 6) from raw where col9 is not null and col9 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id = (select da.id from dataMun_age da where da.from_age = 6));")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col10, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'F%'), (select da.id from dataMun_age da where da.from_age = 6) from raw where col10 is not null and col10 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 6)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col11, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'M%'), (select da.id from dataMun_age da where da.from_age = 10) from raw where col11 is not null and col11 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id=(select da.id from dataMun_age da where da.from_age = 10)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col12, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'F%'), (select da.id from dataMun_age da where da.from_age = 10) from raw where col12 is not null and col12 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 10)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col13, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'M%'), (select da.id from dataMun_age da where da.from_age = 15) from raw where col13 is not null and col13 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id=(select da.id from dataMun_age da where da.from_age = 15)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col14, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'F%'), (select da.id from dataMun_age da where da.from_age = 15) from raw where col14 is not null and col14 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 15)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col15, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'M%'), (select da.id from dataMun_age da where da.from_age = 20) from raw where col15 is not null and col15 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id=(select da.id from dataMun_age da where da.from_age = 20)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col16, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'F%'), (select da.id from dataMun_age da where da.from_age = 20) from raw where col16 is not null and col16 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 20)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col17, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'M%'), (select da.id from dataMun_age da where da.from_age = 55) from raw where col17 is not null and col17 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id=(select da.id from dataMun_age da where da.from_age = 55)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col18, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'F%'), (select da.id from dataMun_age da where da.from_age = 55) from raw where col18 is not null and col18 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 55)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col19, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'M%'), (select da.id from dataMun_age da where da.from_age = 65) from raw where col19 is not null and col19 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id=(select da.id from dataMun_age da where da.from_age = 65)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col20, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'F%'), (select da.id from dataMun_age da where da.from_age = 65) from raw where col20 is not null and col20 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 65)) ;")


def insertWorkbook(spread_sheet):
    start_total = time()
    #conn = psycopg2.connect( host=hostname, user=username, password=password, dbname=database )
    conn = sqlite3.connect('db.sqlite3')
    print(spread_sheet.file)
    start = time()
    print("load_workbook started")
    workbook = openpyxl.load_workbook(filename=spread_sheet.file,read_only=True)
    print(f"load_workbook finished after {round(time()-start,4)} seconds")
    c = conn.cursor()
    start = time()
    print("workbookToSqlStatements started")
    create_sql_statement, insert_sql_statement = workbookToSqlStatements(workbook,"raw","year"," ","week",["col0","col1","col2","col3","col4","col5","col6","col7","col8","col9","col10","col11","col12","col13","col14","col15","col16","col17","col18","col19","col20"],[0,1,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20],70,2)
    print(f"workbookToSqlStatements finished after {round(time()-start,4)} seconds")
    #print("finalizado el srting:",insert_sql_statement)
    #print(create_sql_statement)
    start = time()
    print("create_sql_statement started")
    c.execute(create_sql_statement)
    print(f"create_sql_statement finished after {round(time()-start,4)} seconds")
    delete_table = "delete from raw;"

    start = time()
    print("delete_table started")
    c.execute(delete_table)
    print(f"delete_table finished after {round(time() - start, 4)} seconds")

    start = time()
    print("insert_sql_statement started")
    c.execute(insert_sql_statement)
    print(f"insert_sql_statement finished after {round(time()-start,4)} seconds")
    conn.commit()
    print(f"commit finished after {round(time()-start,4)} seconds")

    start = time()
    print("insert started")
    c.execute("delete from raw where col0 is null;")
    c.execute("insert into dataMun_zone (code) select DISTINCT col0 from raw where col0 not in (select code from dataMun_zone);")
    c.execute("insert into dataMun_center (code,name,zone_id) select DISTINCT col1, col2, (select dz.id from dataMun_zone dz where col0 = dz.code) from raw where col1 not in (select code from dataMun_center);")
    c.execute("insert into dataMun_diagnostic (code,name) select DISTINCT col3, col4 from raw where col3 not in (select code from dataMun_diagnostic);")
    c.execute("insert into dataMun_year (year) select distinct r.year from raw r where r.year not in(select dy.year from dataMun_year dy);")
    c.execute("insert into dataMun_week (week,year_id,creation) select distinct r.week, (SELECT dy.id from dataMun_year dy where dy.year = r.year),current_date from raw r;")
    c.execute("insert or ignore into dataMun_sex(name) values ('Male'),('Female');")
    c.execute("insert or ignore into dataMun_age (from_age,to_age) values(0,1),(1,5),(6,9),(10,14),(15,19),(20,54),(55,65),(65,214748367);")
    """
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date,col5,(select dc.id from dataMun_center dc where r.col1 = dc.code ),(select dd.id from dataMun_diagnostic dd where r.col3 = dd.code),(select dw.id from dataMun_week dw where r.week = dw.week),(select ds.id from dataMun_sex ds where ds.name like 'M%'),(select da.id from dataMun_age da where da.from_age = 0) from raw r where col5 is not null and col5 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id = (select da.id from dataMun_age da where da.from_age = 0));")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col6, (select dc.id from dataMun_center dc where raw.col1 = dc.code ),(select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code),(select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'F%'), (select da.id from dataMun_age da where da.from_age = 0) from raw where col6 is not null and col6 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 0));")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date,col7, (select dc.id from dataMun_center dc where raw.col1 = dc.code ),	(select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code),	(select dw.id from dataMun_week dw where raw.week = dw.week),(select ds.id from dataMun_sex ds where ds.name like 'M%'),	(select da.id from dataMun_age da where da.from_age = 1) from raw where col7 is not null and col7 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id=(select da.id from dataMun_age da where da.from_age = 1));")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date,col8, (select dc.id from dataMun_center dc where raw.col1 = dc.code ),	(select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code),	(select dw.id from dataMun_week dw where raw.week = dw.week),	(select ds.id from dataMun_sex ds where ds.name like 'F%'),	(select da.id from dataMun_age da where da.from_age = 1) from raw where col8 is not null and col8 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 1));")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col9, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'M%'), (select da.id from dataMun_age da where da.from_age = 6) from raw where col9 is not null and col9 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id = (select da.id from dataMun_age da where da.from_age = 6));")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col10, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'F%'), (select da.id from dataMun_age da where da.from_age = 6) from raw where col10 is not null and col10 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 6)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col11, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'M%'), (select da.id from dataMun_age da where da.from_age = 10) from raw where col11 is not null and col11 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id=(select da.id from dataMun_age da where da.from_age = 10)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col12, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'F%'), (select da.id from dataMun_age da where da.from_age = 10) from raw where col12 is not null and col12 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 10)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col13, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'M%'), (select da.id from dataMun_age da where da.from_age = 15) from raw where col13 is not null and col13 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id=(select da.id from dataMun_age da where da.from_age = 15)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col14, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'F%'), (select da.id from dataMun_age da where da.from_age = 15) from raw where col14 is not null and col14 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 15)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col15, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'M%'), (select da.id from dataMun_age da where da.from_age = 20) from raw where col15 is not null and col15 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id=(select da.id from dataMun_age da where da.from_age = 20)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col16, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'F%'), (select da.id from dataMun_age da where da.from_age = 20) from raw where col16 is not null and col16 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 20)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col17, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'M%'), (select da.id from dataMun_age da where da.from_age = 55) from raw where col17 is not null and col17 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id=(select da.id from dataMun_age da where da.from_age = 55)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col18, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'F%'), (select da.id from dataMun_age da where da.from_age = 55) from raw where col18 is not null and col18 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 55)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col19, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'M%'), (select da.id from dataMun_age da where da.from_age = 65) from raw where col19 is not null and col19 not in (select cases from dataMun_diagnosticcases where sex_id=1 and age_id=(select da.id from dataMun_age da where da.from_age = 65)) ;")
    c.execute("insert into dataMun_diagnosticcases (creation, cases, center_id, diagnostic_id, week_id, sex_id, age_id) select current_date, col20, (select dc.id from dataMun_center dc where raw.col1 = dc.code ), (select dd.id from dataMun_diagnostic dd where raw.col3 = dd.code), (select dw.id from dataMun_week dw where raw.week = dw.week), (select ds.id from dataMun_sex ds where ds.name like 'F%'), (select da.id from dataMun_age da where da.from_age = 65) from raw where col20 is not null and col20 not in (select cases from dataMun_diagnosticcases where sex_id=2 and age_id=(select da.id from dataMun_age da where da.from_age = 65)) ;")
    """
    conn.commit()
    print(f"insert finish after {round(time() - start, 4)} seconds")

    workbook.close()
    conn.close()
    print(f"total finished after {round(time()-start_total,4)} seconds")

class DotsGraphicAverage():
    def __init__(self, average, week, lower_rank, top_rank,cases):
        self.average = average
        self.week = week
        self.lower_rank = lower_rank
        self.top_rank = top_rank
        self.cases = cases



def GetGraphicAverages(diagnostic_cases, diagnostic, weeks,year, n_years):
    t = [4.30, 3.18, 2.78, 2.57, 2.45, 2.36, 2.31, 2.26, 2.23, 2.20]

    current_year = Year.objects.get(year=year)
    weeks_current_year = weeks.filter(year=current_year)
    year_ob = Year.objects.filter(year__lt=year)
    weeks = weeks.filter(year__in=year_ob)

    population = [0] * n_years
    popu = 0

    #cases per diagnostic
    diagnostic_cases_w = diagnostic_cases

    #arithmetic average of the weeks / n_years
    averages = [0] * 52

    standard_deviations = [0] * 52
    #number of years

    #cases per week of the diferent years
    cases_per_weeks = [0] * 52

    for i in range(len(averages)):

        f = [0]*(n_years-1)

        year = 0

        y = 0
        for w in range(len(weeks)):
            #print(y)
            if weeks[w].week == i+1:
                if y == 0:
                    year = weeks[w].year
                if year != weeks[w].year: # Esto no pasa nunca
                    year = weeks[w].year

                cases = 0
                pop = weeks[w].year.population
                population[y] = pop
                for p in diagnostic_cases_w:

                    if p.week == weeks[w]:

                        cases += p.cases

                f[y] = (cases / population[y] * 100000)
                y+=1


        averages[i] = np.log(np.average(f))

        standard_deviations[i] = np.std(f)


        """
            if cases != 0:

                average = cases / n_years

                averages[i] = average
                #calculation of standar deviation
                standard_deviation = 0 
                if len(f) != 0:
                    for cases in f:
                        suma2 += (cases-average)**2

                    standard_deviation = math.sqrt(suma2 / len(f))
                    standard_deviations[i] = standard_deviation 
        """
            
        cases = 0
        for week in weeks_current_year:
            if week.week == i+1:
                dia = diagnostic_cases.filter(week=week)
                
                for d in dia:

                    cases += d.cases
            popu = week.year.population

        cases_per_weeks[i] = np.log(cases/popu*100000)




    #array of class dots for draw the chart of averages
    dots_graphic_averages = []
    #array of class dots for draw the chart of cumulative
    dots_graphic_cumulative = []


    average_cumulative = 0
    top_rank_cumulative = 0
    cases_acumulative = 0
    lower_rank_cumulative = 0

    for i in range(len(standard_deviations)):
        lower_rank = 0
        top_rank = 0
        averages[i] = np.exp(averages[i])
        averages[i] = averages[i] * popu / 100000

        standard_deviations[i] = standard_deviations[i]

        standard_deviations[i] = standard_deviations[i] * popu / 100000


        if n_years != 0:
            lower_rank = averages[i] - (t[n_years-3] * standard_deviations[i]/ math.sqrt(n_years))
            top_rank = averages[i] + (t[n_years-3] * standard_deviations[i] / math.sqrt(n_years))
            if lower_rank < 0:
                lower_rank = 0


        cases_per_weeks[i] = np.exp(cases_per_weeks[i])
        cases_per_weeks[i] = cases_per_weeks[i] * popu / 100000

        # Acumulative dots
        cases_acumulative += cases_per_weeks[i]
        average_cumulative += averages[i]
        if lower_rank >= 0:
            lower_rank_cumulative += lower_rank
        top_rank_cumulative += top_rank



        dots_average = DotsGraphicAverage(averages[i],i+1, lower_rank, top_rank,cases_per_weeks[i])
        dots_cumulative = DotsGraphicAverage(average_cumulative,i+1, lower_rank_cumulative, top_rank_cumulative,cases_acumulative)
        dots_graphic_averages.append(dots_average)
        dots_graphic_cumulative.append(dots_cumulative)


    return dots_graphic_averages, dots_graphic_cumulative


def GetAlert(diagnostic_cases, diagnostic, week,year):

    #cases per diagnostic

    diag_cases = diagnostic_cases.filter(diagnostic=diagnostic)
    average = 0
    standard_deviation = 0
    cases = 0
    #number of years
    n_years = 0
    year_var = 0
    f = []
    weeks = Week.objects.filter(year__lt=year,week=week.week).order_by('year')
    for w in weeks:


        if year_var != w.year:
            n_years += 1
            year_var = w.year


        pac = diag_cases.filter(week=w)
        x = 0
        for p in pac:

            cases += p.cases
            x = p.cases
            f.append(x)

    if cases != 0:

        average = cases / n_years

        #calculation of standar deviation
        if len(f) != 1:
            suma2 = 0
            for cases in f:
                suma2 += (cases-average)**2
            standard_deviation = math.sqrt(suma2 / len(f))
    cases = 0
    dia = diag_cases.filter(week=week)

    for d in dia:
        cases += d.cases

    #array of class dots for draw the chart

    lower_rank = 0
    top_rank = 0
    if n_years != 0:
        lower_rank = average - (3.18 * standard_deviation / math.sqrt(n_years))
        top_rank = average + (3.18 * standard_deviation / math.sqrt(n_years))

    dots = DotsGraphicAverage(average,week.week, lower_rank, top_rank,cases)

    return dots


class DotsGraphicQuartile():
    def __init__(self,q1,q2,q3,cases,week):
        self.q1 = q1
        self.q2 = q2
        self.q3 = q3
        self.cases = cases
        self.week = week


def GetGraphicQuartiles(diagnostic_cases, diagnostic, weeks,year, n_years):
    #idk la verdad
    current_year = Year.objects.get(year=year)
    weeks_current_year = weeks.filter(year=current_year)
    year_ob = Year.objects.filter(year__lt=year)
    weeks = weeks.filter(year__in=year_ob)
    print(weeks)


    suma = 0
    if n_years % 2 != 0:
        suma = -1
    q = [(n_years+suma)*1 / 4,(n_years+suma)*2 / 4,(n_years+suma)*3 / 4]
    qs = [0] *3


    diagnost_cases = diagnostic_cases


    dots_graphic_quartiles = [ ]
    for o in range(52):
        dots_q=[]

        cases_per_years = [0] * (n_years-1)


        cases = 0
        i = 0
        for week in range(len(weeks)):
            if weeks[week].week == o+1:
                cases = 0
                for p in diagnostic_cases:
                    if p.week == weeks[week]:
                        cases += p.cases
                cases_per_years[i] = cases
                i += 1

        qs[0] = np.quantile(cases_per_years, 0.25)
        qs[1] = np.quantile(cases_per_years, 0.5)
        qs[2] = np.quantile(cases_per_years, 0.75)
        print(qs)
        print(cases_per_years)

        cases_per_week = 0

        for week in weeks_current_year:
            if week.week == o+1 :
                dia = diagnost_cases.filter(week=week)

                suma = 0
                for d in dia:
                    cases_per_week += d.cases


        #cases_per_week = suma
        #print(cases_per_week)

        dots = DotsGraphicQuartile(qs[0],qs[1],qs[2],cases_per_week,o+1)
        dots_graphic_quartiles.append(dots)

    return dots_graphic_quartiles


